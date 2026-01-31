from __future__ import annotations

import base64
import dataclasses
import datetime as dt
import hashlib
import io
import json
import re
import tempfile
import threading
import time
import uuid
from concurrent.futures import Future, ThreadPoolExecutor
from dataclasses import dataclass
from typing import Any, Callable, Literal, Optional

import pandas as pd
import plotly.graph_objects as go
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet
from plotly.offline import get_plotlyjs

ExportFormat = Literal["html", "xlsx"]


@dataclass(frozen=True)
class CardSummary:
    name: str
    total: Optional[float]
    average: Optional[float]


@dataclass(frozen=True)
class ChartBundle:
    title: str
    table: pd.DataFrame
    figure: Optional[go.Figure] = None
    echarts_option: Optional[dict[str, Any]] = None
    echarts_extra: Optional[dict[str, Any]] = None


@dataclass
class ExportEstimate:
    estimate_seconds: float
    size_bytes: int


@dataclass
class ExportResult:
    content: bytes
    sha256: str


@dataclass
class ExportTask:
    task_id: str
    user_email: str
    export_format: ExportFormat
    progress: int = 0
    status: Literal["queued", "running", "succeeded", "failed", "cancelled"] = "queued"
    trace_id: Optional[str] = None
    error_message: Optional[str] = None
    estimate_seconds: Optional[float] = None
    size_bytes: Optional[int] = None
    sha256: Optional[str] = None
    created_at_utc: str = dataclasses.field(default_factory=lambda: dt.datetime.now(dt.timezone.utc).isoformat())
    finished_at_utc: Optional[str] = None
    cancelled: bool = False
    _future: Optional[Future[ExportResult]] = None


def _now_utc_iso() -> str:
    return dt.datetime.now(dt.timezone.utc).isoformat()


def _sha256(data: bytes) -> str:
    return hashlib.sha256(data).hexdigest()


def _safe_sheet_title(title: str) -> str:
    title = re.sub(r"[\\/*?:\[\]]", "_", title)
    title = title.strip() or "Sheet"
    return title[:31]


def format_number(value: Any) -> str:
    try:
        v = float(value)
    except Exception:
        return str(value)
    if pd.isna(v):
        return "—"
    if abs(v) >= 1_0000_0000:
        return f"{v/1_0000_0000:.2f}亿"
    if abs(v) >= 1_0000:
        return f"{v/1_0000:.2f}万"
    return f"{v:,.2f}"


def estimate_export(num_cards: int, num_charts: int, export_format: ExportFormat) -> ExportEstimate:
    if export_format == "html":
        base = 3_700_000
        per_chart = 25_000
        per_card = 450
        size_bytes = base + num_charts * per_chart + num_cards * per_card
        estimate_seconds = 0.25 + num_charts * 0.12 + num_cards * 0.01
        return ExportEstimate(estimate_seconds=estimate_seconds, size_bytes=size_bytes)

    base = 40_000
    per_chart = 220_000
    per_card = 6_000
    size_bytes = base + num_charts * per_chart + num_cards * per_card
    estimate_seconds = 0.5 + num_charts * 0.35 + num_cards * 0.03
    return ExportEstimate(estimate_seconds=estimate_seconds, size_bytes=size_bytes)


def build_card_summaries(df: pd.DataFrame, card_metrics: list[str]) -> list[CardSummary]:
    cards: list[CardSummary] = []
    for m in card_metrics:
        if m not in df.columns:
            continue
        try:
            total = float(df[m].sum())
        except Exception:
            total = None
        try:
            avg = float(df[m].mean())
        except Exception:
            avg = None
        cards.append(CardSummary(name=m, total=total, average=avg))
    return cards


def _extract_x(df: pd.DataFrame, x_axis_mode: str) -> tuple[pd.Series, str]:
    if "时间列:" in str(x_axis_mode):
        time_col = str(x_axis_mode).split(": ", 1)[1]
        x = pd.to_datetime(df[time_col], errors="coerce")
        x = x.dt.strftime("%Y-%m-%d").fillna(df[time_col].astype(str))
        return x, time_col
    if "类别列:" in str(x_axis_mode):
        cat_col = str(x_axis_mode).split(": ", 1)[1]
        return df[cat_col].astype(str), cat_col
    return pd.Series(range(len(df))), "序号"


def _echarts_extract_x(df: pd.DataFrame, x_axis_mode: str) -> tuple[pd.DataFrame, list[Any], str]:
    if "时间列:" in str(x_axis_mode):
        time_col = str(x_axis_mode).split(": ", 1)[1]
        tmp = df.copy()
        if time_col in tmp.columns and pd.api.types.is_datetime64_any_dtype(tmp[time_col]):
            x_data = tmp[time_col].dt.strftime("%Y-%m-%d").tolist()
        else:
            x_data = tmp[time_col].astype(str).tolist() if time_col in tmp.columns else list(range(len(tmp)))
        return tmp.reset_index(drop=True), x_data, time_col
    if "类别列:" in str(x_axis_mode):
        cat_col = str(x_axis_mode).split(": ", 1)[1]
        use_cols = [c for c in df.columns if c != cat_col]
        tmp = df.groupby(cat_col)[use_cols].sum(numeric_only=True).reset_index()
        x_data = tmp[cat_col].astype(str).tolist()
        return tmp, x_data, cat_col
    tmp = df.reset_index(drop=True)
    return tmp, tmp.index.tolist(), "序号"


def _geojson_feature_name(feature: dict) -> str:
    props = feature.get("properties") or {}
    for k in ("name", "NAME", "Name", "fullname", "FULLNAME", "fullName"):
        v = props.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _geojson_feature_centroid(feature: dict) -> Optional[tuple[float, float]]:
    geom = feature.get("geometry") or {}
    gtype = geom.get("type")
    coords = geom.get("coordinates")
    if not coords:
        return None

    pts: list[tuple[float, float]] = []
    try:
        if gtype == "Polygon":
            rings = coords
            for ring in rings:
                for p in ring:
                    if isinstance(p, (list, tuple)) and len(p) >= 2:
                        pts.append((float(p[0]), float(p[1])))
        elif gtype == "MultiPolygon":
            for poly in coords:
                for ring in poly:
                    for p in ring:
                        if isinstance(p, (list, tuple)) and len(p) >= 2:
                            pts.append((float(p[0]), float(p[1])))
    except Exception:
        return None

    if not pts:
        return None
    xs = [p[0] for p in pts]
    ys = [p[1] for p in pts]
    return (sum(xs) / len(xs), sum(ys) / len(ys))


def _build_echarts_bundle(df: pd.DataFrame, conf: dict[str, Any]) -> Optional[ChartBundle]:
    ctype = conf.get("type")
    metrics: list[str] = conf.get("metrics") or []
    dimension = conf.get("dimension")
    x_axis_mode = conf.get("x_axis_mode")
    style = str(conf.get("chart_style") or "默认")
    dimension2 = str(conf.get("dimension2") or "不选择")
    target_value = float(conf.get("target_value") or 0)
    agg_mode = str(conf.get("agg_mode") or "汇总")
    radar_topn = int(conf.get("radar_topn") or 5)
    map_geojson = conf.get("map_geojson")
    lon_col = str(conf.get("lon_col") or "不选择")
    lat_col = str(conf.get("lat_col") or "不选择")

    colors = ["#1277D1", "#2FA3FF", "#FF8A3D", "#FF5252", "#8C52FF", "#66B2FF"]

    if ctype == "趋势图":
        if not metrics:
            return None
        tmp, x_data, x_name = _echarts_extract_x(df, str(x_axis_mode))
        series = []
        for idx, metric in enumerate(metrics):
            if metric not in tmp.columns:
                continue
            area_style = None
            line_style = {"width": 2}
            item_style = {"color": colors[idx % len(colors)]}
            if style == "面积":
                area_style = {
                    "color": {
                        "type": "linear",
                        "x": 0,
                        "y": 0,
                        "x2": 0,
                        "y2": 1,
                        "colorStops": [
                            {"offset": 0, "color": f"{colors[idx % len(colors)]}88"},
                            {"offset": 1, "color": f"{colors[idx % len(colors)]}00"},
                        ],
                    }
                }
            if style == "大屏发光":
                line_style = {"width": 3, "shadowBlur": 12, "shadowColor": colors[idx % len(colors)]}
                item_style = {
                    "color": colors[idx % len(colors)],
                    "shadowBlur": 12,
                    "shadowColor": colors[idx % len(colors)],
                }
            series.append(
                {
                    "name": metric,
                    "type": "line",
                    "smooth": True,
                    "showSymbol": False,
                    "data": tmp[metric].tolist(),
                    "itemStyle": item_style,
                    "lineStyle": line_style,
                    "areaStyle": area_style,
                }
            )
        series = [s for s in series if s.get("data")]
        option = {
            "title": {"text": "趋势分析", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "axis"},
            "legend": {"data": [s["name"] for s in series], "textStyle": {"color": "#3A4A63"}, "top": 25},
            "grid": {"left": "3%", "right": "4%", "bottom": "3%", "containLabel": True},
            "xAxis": {
                "type": "category",
                "boundaryGap": False,
                "data": x_data,
                "axisLabel": {"color": "#3A4A63"},
                "axisLine": {"lineStyle": {"color": "#A7C9FF"}},
            },
            "yAxis": {
                "type": "value",
                "splitLine": {"lineStyle": {"color": "rgba(167, 201, 255, 0.3)", "type": "dashed"}},
                "axisLabel": {"color": "#3A4A63"},
            },
            "series": series,
        }
        title = " · ".join([s["name"] for s in series]) + " 趋势" if series else "趋势图"
        table = tmp[[x_name] + [s["name"] for s in series]] if x_name in tmp.columns else tmp[[s["name"] for s in series]]
        return ChartBundle(title=title, table=table, echarts_option=option)

    if ctype == "占比图":
        if not metrics or not dimension or dimension == "不选择" or metrics[0] not in df.columns or dimension not in df.columns:
            return None
        metric = metrics[0]
        pie_data = df.groupby(dimension)[metric].sum(numeric_only=True).reset_index().sort_values(by=metric, ascending=False)
        pie_chart_data = [{"value": float(row[metric] or 0), "name": str(row[dimension])} for _, row in pie_data.iterrows()]

        rose = None
        radius: Any = ["40%", "70%"]
        if style == "玫瑰":
            rose = "radius"
            radius = ["20%", "70%"]
        if style == "饼图":
            rose = None
            radius = "65%"

        option = {
            "title": {"text": f"{metric} 占比", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "item"},
            "legend": {
                "orient": "vertical",
                "left": "left",
                "textStyle": {"color": "#3A4A63"},
                "top": 25,
            },
            "series": [
                {
                    "name": metric,
                    "type": "pie",
                    "radius": radius,
                    "roseType": rose,
                    "avoidLabelOverlap": False,
                    "itemStyle": {"borderRadius": 5, "borderColor": "#fff", "borderWidth": 2},
                    "label": {"show": False, "position": "center"},
                    "emphasis": {"label": {"show": True, "fontSize": "18", "fontWeight": "bold"}},
                    "labelLine": {"show": False},
                    "data": pie_chart_data,
                }
            ],
            "color": ["#1277D1", "#2FA3FF", "#66B2FF", "#99CCFF", "#CCE5FF", "#FF8A3D"],
        }
        title = f"{metric} 按 {dimension} 占比"
        return ChartBundle(title=title, table=pie_data, echarts_option=option)

    if ctype == "排名图":
        if not metrics or not dimension or dimension == "不选择" or metrics[0] not in df.columns or dimension not in df.columns:
            return None
        metric = metrics[0]
        bar_data = df.groupby(dimension)[metric].sum(numeric_only=True).reset_index()
        top_data = bar_data.sort_values(by=metric, ascending=True).tail(10)
        categories = top_data[dimension].astype(str).tolist()
        values = top_data[metric].tolist()

        if style == "竖向":
            axis_series = {
                "xAxis": {"type": "category", "data": categories, "axisLabel": {"color": "#3A4A63"}},
                "yAxis": {
                    "type": "value",
                    "axisLabel": {"color": "#3A4A63"},
                    "splitLine": {"lineStyle": {"color": "rgba(167, 201, 255, 0.3)", "type": "dashed"}},
                },
            }
        else:
            axis_series = {
                "xAxis": {
                    "type": "value",
                    "axisLabel": {"color": "#3A4A63"},
                    "splitLine": {"lineStyle": {"color": "rgba(167, 201, 255, 0.3)", "type": "dashed"}},
                },
                "yAxis": {"type": "category", "data": categories, "axisLabel": {"color": "#3A4A63"}},
            }

        realtime_sort = style == "实时排名赛" and style != "竖向"

        option = {
            "title": {"text": f"{metric} Top 10", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "axis", "axisPointer": {"type": "shadow"}},
            "grid": {"left": "3%", "right": "4%", "bottom": "3%", "containLabel": True},
            **axis_series,
            "animationDuration": 800,
            "animationDurationUpdate": 800,
            "series": [
                {
                    "name": metric,
                    "type": "bar",
                    "data": values,
                    "realtimeSort": realtime_sort,
                    "label": {"show": realtime_sort, "position": "right", "color": "#0B1B33"},
                    "itemStyle": {
                        "color": {
                            "type": "linear",
                            "x": 0,
                            "y": 0,
                            "x2": 1,
                            "y2": 0,
                            "colorStops": [{"offset": 0, "color": "#2FA3FF"}, {"offset": 1, "color": "#1277D1"}],
                        },
                        "borderRadius": [0, 4, 4, 0],
                    },
                }
            ],
        }
        title = f"{metric} 按 {dimension} 排名"
        return ChartBundle(title=title, table=top_data, echarts_option=option)

    if ctype == "漏斗图":
        if not metrics:
            return None
        data: list[dict[str, Any]] = []
        title_text = "转化漏斗"
        if dimension and dimension != "不选择" and metrics:
            metric = metrics[0]
            if metric not in df.columns or dimension not in df.columns:
                return None
            funnel_data = df.groupby(dimension)[metric].sum(numeric_only=True).reset_index().sort_values(by=metric, ascending=False)
            data = [{"value": float(row[metric] or 0), "name": str(row[dimension])} for _, row in funnel_data.iterrows()]
            title_text = f"{metric} 按 {dimension} 漏斗"
            table = funnel_data
        elif len(metrics) > 1:
            use_cols = [m for m in metrics if m in df.columns]
            sums = df[use_cols].sum(numeric_only=True).sort_values(ascending=False)
            data = [{"value": float(val or 0), "name": str(name)} for name, val in sums.items()]
            title_text = "多指标转化漏斗"
            table = pd.DataFrame({"stage": [d["name"] for d in data], "value": [d["value"] for d in data]})
        else:
            return None

        option = {
            "title": {"text": title_text, "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "item", "formatter": "{a} <br/>{b} : {c}"},
            "legend": {"data": [d["name"] for d in data], "textStyle": {"color": "#3A4A63"}, "top": 25},
            "series": [
                {
                    "name": "漏斗",
                    "type": "funnel",
                    "left": "10%",
                    "top": 60,
                    "bottom": 60,
                    "width": "80%",
                    "min": 0,
                    "max": data[0]["value"] if data else 100,
                    "minSize": "0%",
                    "maxSize": "100%",
                    "sort": "descending",
                    "gap": 2,
                    "label": {"show": True, "position": "inside"},
                    "labelLine": {"length": 10, "lineStyle": {"width": 1, "type": "solid"}},
                    "itemStyle": {"borderColor": "#fff", "borderWidth": 1},
                    "emphasis": {"label": {"fontSize": 20}},
                    "data": data,
                }
            ],
            "color": ["#1277D1", "#2FA3FF", "#66B2FF", "#99CCFF", "#FF8A3D"],
        }
        return ChartBundle(title=title_text, table=table, echarts_option=option)

    if ctype == "箱线图":
        if not metrics or not dimension or dimension == "不选择" or metrics[0] not in df.columns or dimension not in df.columns:
            return None
        metric = metrics[0]
        grouped = df[[dimension, metric]].dropna().copy()
        grouped[dimension] = grouped[dimension].astype(str)
        medians = grouped.groupby(dimension)[metric].median().sort_values(ascending=False)
        cats = medians.index.tolist()[:20]
        grouped = grouped[grouped[dimension].isin(cats)]

        data = []
        for c in cats:
            values = grouped[grouped[dimension] == c][metric]
            if len(values) == 0:
                continue
            q1 = float(values.quantile(0.25))
            q2 = float(values.quantile(0.50))
            q3 = float(values.quantile(0.75))
            vmin = float(values.min())
            vmax = float(values.max())
            data.append([vmin, q1, q2, q3, vmax])

        option = {
            "title": {"text": f"{metric} 分布（箱线）", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "item"},
            "grid": {"left": "3%", "right": "4%", "bottom": "10%", "containLabel": True},
            "xAxis": {"type": "category", "data": cats, "axisLabel": {"color": "#3A4A63", "rotate": 25}},
            "yAxis": {
                "type": "value",
                "axisLabel": {"color": "#3A4A63"},
                "splitLine": {"lineStyle": {"color": "rgba(167, 201, 255, 0.3)", "type": "dashed"}},
            },
            "series": [
                {"name": metric, "type": "boxplot", "data": data, "itemStyle": {"color": "#2FA3FF", "borderColor": "#1277D1"}}
            ],
        }
        return ChartBundle(title=f"{metric} 按 {dimension} 分布（箱线）", table=grouped, echarts_option=option)

    if ctype == "散点图":
        if len(metrics) < 2:
            return None
        x_metric, y_metric = metrics[0], metrics[1]
        if x_metric not in df.columns or y_metric not in df.columns:
            return None
        series_type = "effectScatter" if style == "涟漪" else "scatter"
        tmp = df[[x_metric, y_metric] + ([dimension] if dimension and dimension != "不选择" and dimension in df.columns else [])].dropna().copy()
        tmp[x_metric] = pd.to_numeric(tmp[x_metric], errors="coerce")
        tmp[y_metric] = pd.to_numeric(tmp[y_metric], errors="coerce")
        tmp = tmp.dropna(subset=[x_metric, y_metric]).head(2000)

        series = []
        if dimension and dimension != "不选择" and dimension in tmp.columns:
            top_cats = tmp[dimension].astype(str).value_counts().head(8).index.tolist()
            for idx, c in enumerate(top_cats):
                sub = tmp[tmp[dimension].astype(str) == c]
                series.append(
                    {
                        "name": str(c),
                        "type": series_type,
                        "data": sub[[x_metric, y_metric]].values.tolist(),
                        "symbolSize": 9,
                        "itemStyle": {"color": colors[idx % len(colors)]},
                        "rippleEffect": {"scale": 3, "brushType": "stroke"} if series_type == "effectScatter" else None,
                    }
                )
        else:
            series.append(
                {
                    "name": f"{x_metric} vs {y_metric}",
                    "type": series_type,
                    "data": tmp[[x_metric, y_metric]].values.tolist(),
                    "symbolSize": 9,
                    "itemStyle": {"color": "#1277D1"},
                    "rippleEffect": {"scale": 3, "brushType": "stroke"} if series_type == "effectScatter" else None,
                }
            )

        series = [s for s in series if s.get("data")]
        for s in series:
            if s.get("rippleEffect") is None:
                s.pop("rippleEffect", None)

        option = {
            "title": {"text": f"{x_metric} vs {y_metric}", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "item"},
            "legend": {"show": len(series) > 1, "textStyle": {"color": "#3A4A63"}, "top": 25},
            "grid": {"left": "3%", "right": "4%", "bottom": "8%", "containLabel": True},
            "xAxis": {"type": "value", "name": x_metric, "axisLabel": {"color": "#3A4A63"}},
            "yAxis": {"type": "value", "name": y_metric, "axisLabel": {"color": "#3A4A63"}},
            "series": series,
        }
        return ChartBundle(title=f"{x_metric} vs {y_metric} 散点", table=tmp, echarts_option=option)

    if ctype == "热力图":
        if not metrics or not dimension or dimension == "不选择" or dimension2 == "不选择":
            return None
        metric = metrics[0]
        if metric not in df.columns or dimension not in df.columns or dimension2 not in df.columns:
            return None
        tmp = df[[dimension, dimension2, metric]].dropna().copy()
        tmp[dimension] = tmp[dimension].astype(str)
        tmp[dimension2] = tmp[dimension2].astype(str)

        top_x = tmp.groupby(dimension)[metric].sum(numeric_only=True).sort_values(ascending=False).head(20).index.tolist()
        top_y = tmp.groupby(dimension2)[metric].sum(numeric_only=True).sort_values(ascending=False).head(20).index.tolist()
        tmp = tmp[tmp[dimension].isin(top_x) & tmp[dimension2].isin(top_y)]

        pivot = tmp.groupby([dimension2, dimension])[metric].sum(numeric_only=True).reset_index()
        x_labels = top_x
        y_labels = top_y
        x_index = {v: i for i, v in enumerate(x_labels)}
        y_index = {v: i for i, v in enumerate(y_labels)}

        data = []
        vmax = 0.0
        for _, row in pivot.iterrows():
            x = x_index.get(str(row[dimension]))
            y = y_index.get(str(row[dimension2]))
            v = float(row[metric] or 0)
            vmax = max(vmax, v)
            if x is not None and y is not None:
                data.append([x, y, v])

        option = {
            "title": {"text": f"{metric} 热力分布", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"position": "top"},
            "grid": {"left": "7%", "right": "4%", "bottom": "12%", "containLabel": True},
            "xAxis": {"type": "category", "data": x_labels, "axisLabel": {"color": "#3A4A63", "rotate": 25}},
            "yAxis": {"type": "category", "data": y_labels, "axisLabel": {"color": "#3A4A63"}},
            "visualMap": {
                "min": 0,
                "max": max(1.0, vmax),
                "calculable": True,
                "orient": "horizontal",
                "left": "center",
                "bottom": 0,
                "inRange": {"color": ["#CCE5FF", "#2FA3FF", "#1277D1"]},
            },
            "series": [
                {
                    "name": metric,
                    "type": "heatmap",
                    "data": data,
                    "label": {"show": False},
                    "emphasis": {"itemStyle": {"shadowBlur": 10, "shadowColor": "rgba(0, 0, 0, 0.25)"}},
                }
            ],
        }
        return ChartBundle(title=f"{metric} 在 {dimension}×{dimension2} 热力", table=pivot, echarts_option=option)

    if ctype == "雷达图":
        if not metrics or not dimension or dimension == "不选择":
            return None
        use_metrics = [m for m in metrics[:6] if m in df.columns]
        if not use_metrics or dimension not in df.columns:
            return None
        tmp = df[[dimension] + use_metrics].dropna().copy()
        tmp[dimension] = tmp[dimension].astype(str)
        g = tmp.groupby(dimension)[use_metrics].sum(numeric_only=True).reset_index()
        sort_metric = use_metrics[0]
        g = g.sort_values(by=sort_metric, ascending=False).head(max(1, int(radar_topn or 5)))

        indicators = []
        for m in use_metrics:
            vmax = float(g[m].max() or 0)
            indicators.append({"name": m, "max": max(1.0, vmax * 1.2)})

        series_data = []
        for _, row in g.iterrows():
            series_data.append({"name": str(row[dimension]), "value": [float(row[m] or 0) for m in use_metrics]})
        if style == "单体" and series_data:
            series_data = [series_data[0]]

        option = {
            "title": {"text": "多维指标对比", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {},
            "legend": {"data": [d["name"] for d in series_data], "textStyle": {"color": "#3A4A63"}, "top": 25},
            "radar": {
                "indicator": indicators,
                "splitArea": {"areaStyle": {"color": ["rgba(18, 119, 209, 0.06)"]}},
                "axisLine": {"lineStyle": {"color": "rgba(167, 201, 255, 0.7)"}},
                "splitLine": {"lineStyle": {"color": "rgba(167, 201, 255, 0.45)"}},
            },
            "series": [{"type": "radar", "data": series_data, "areaStyle": {"opacity": 0.12}}],
            "color": ["#1277D1", "#2FA3FF", "#FF8A3D", "#8C52FF", "#FF5252"],
        }
        return ChartBundle(title=f"{dimension} 多指标雷达", table=g, echarts_option=option)

    if ctype == "仪表盘":
        if not metrics:
            return None
        metric = metrics[0]
        if metric not in df.columns:
            return None
        series = pd.to_numeric(df[metric], errors="coerce")
        v = float(series.mean() if str(agg_mode) == "均值" else series.sum())
        if pd.isna(v):
            v = 0.0
        max_v = float(target_value or 0)
        if max_v <= 0:
            max_v = max(1.0, abs(v) * 1.2)
        start_angle, end_angle = (180, 0) if style == "半圆" else (90, -270)
        ratio = min(max(v / max_v, 0.0), 1.0)

        option = {
            "title": {"text": f"{metric} 完成度", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "series": [
                {
                    "type": "gauge",
                    "startAngle": start_angle,
                    "endAngle": end_angle,
                    "min": 0,
                    "max": max_v,
                    "progress": {"show": True, "width": 16},
                    "axisLine": {"lineStyle": {"width": 16, "color": [[ratio, "#1277D1"], [1, "#CCE5FF"]]}},
                    "axisTick": {"show": False},
                    "splitLine": {"show": False},
                    "axisLabel": {"color": "#3A4A63"},
                    "pointer": {"show": True, "width": 4},
                    "title": {"show": True, "offsetCenter": [0, "65%"], "color": "#3A4A63"},
                    "detail": {"valueAnimation": True, "formatter": "{value}", "color": "#0B1B33"},
                    "data": [{"value": round(v, 4), "name": str(agg_mode or "汇总")}],
                }
            ],
        }
        table = pd.DataFrame({"metric": [metric], "value": [v], "target": [max_v], "agg_mode": [agg_mode]})
        return ChartBundle(title=f"{metric} 仪表盘", table=table, echarts_option=option)

    if ctype == "桑基图":
        if not metrics or not dimension or dimension == "不选择" or dimension2 == "不选择":
            return None
        metric = metrics[0]
        if metric not in df.columns or dimension not in df.columns or dimension2 not in df.columns:
            return None
        tmp = df[[dimension, dimension2, metric]].dropna().copy()
        tmp[dimension] = tmp[dimension].astype(str)
        tmp[dimension2] = tmp[dimension2].astype(str)
        g = tmp.groupby([dimension, dimension2])[metric].sum(numeric_only=True).reset_index()
        g = g.sort_values(by=metric, ascending=False).head(40)
        nodes = sorted(set(g[dimension].astype(str).tolist() + g[dimension2].astype(str).tolist()))
        links = [
            {"source": str(r[dimension]), "target": str(r[dimension2]), "value": float(r[metric] or 0)}
            for _, r in g.iterrows()
        ]
        option = {
            "title": {"text": f"{metric} 流向", "textStyle": {"color": "#1277D1", "fontSize": 16}},
            "tooltip": {"trigger": "item", "triggerOn": "mousemove"},
            "series": [
                {
                    "type": "sankey",
                    "data": [{"name": n} for n in nodes],
                    "links": links,
                    "emphasis": {"focus": "adjacency"},
                    "lineStyle": {"color": "source", "curveness": 0.5, "opacity": 0.45},
                    "label": {"color": "#0B1B33"},
                }
            ],
            "color": ["#1277D1", "#2FA3FF", "#FF8A3D", "#8C52FF", "#FF5252", "#66B2FF"],
        }
        return ChartBundle(title=f"{metric} 从 {dimension} 到 {dimension2} 流向", table=g, echarts_option=option)

    if ctype == "地图":
        if not metrics:
            return None
        metric = metrics[0]
        if metric not in df.columns:
            return None

        if isinstance(map_geojson, dict) and map_geojson.get("type") == "FeatureCollection" and map_geojson.get("features"):
            map_name = "CUSTOM_MAP"
            features = map_geojson.get("features") or []
            name_to_centroid: dict[str, tuple[float, float]] = {}
            for f in features:
                n = _geojson_feature_name(f)
                c = _geojson_feature_centroid(f)
                if n and c:
                    name_to_centroid[n] = c

            if dimension and dimension != "不选择" and dimension in df.columns:
                tmp = df[[dimension, metric]].dropna().copy()
                tmp[dimension] = tmp[dimension].astype(str)
                tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
                tmp = tmp.dropna(subset=[metric])
                g = tmp.groupby(dimension)[metric].sum(numeric_only=True).reset_index()
            else:
                g = pd.DataFrame({dimension: ["总计"], metric: [pd.to_numeric(df[metric], errors="coerce").sum()]})

            g = g.sort_values(by=metric, ascending=False)
            vmax = float(g[metric].max() or 0)
            if vmax <= 0:
                vmax = 1.0

            base_geo = {
                "map": map_name,
                "roam": True,
                "label": {"show": False},
                "itemStyle": {"areaColor": "#E6F3FF", "borderColor": "#A7C9FF"},
                "emphasis": {"itemStyle": {"areaColor": "#CCE5FF"}},
            }

            series_out: list[dict[str, Any]] = []
            if style == "区域填色":
                series_out.append(
                    {
                        "type": "map",
                        "map": map_name,
                        "geoIndex": 0,
                        "data": [
                            {"name": str(r[dimension]), "value": float(r[metric] or 0)}
                            for _, r in g.iterrows()
                            if str(r[dimension])
                        ],
                    }
                )

            if style in {"散点叠加", "航线叠加"}:
                points = []
                for _, r in g.head(80).iterrows():
                    name = str(r[dimension])
                    c = name_to_centroid.get(name)
                    if not c:
                        continue
                    points.append({"name": name, "value": [c[0], c[1], float(r[metric] or 0)]})
                if points:
                    series_out.append(
                        {
                            "type": "effectScatter" if style == "散点叠加" else "scatter",
                            "coordinateSystem": "geo",
                            "symbolSize": 10,
                            "data": points,
                            "itemStyle": {"color": "#1277D1"},
                            "rippleEffect": {"scale": 3, "brushType": "stroke"} if style == "散点叠加" else None,
                        }
                    )

            if (
                style == "航线叠加"
                and dimension
                and dimension != "不选择"
                and dimension2
                and dimension2 != "不选择"
                and dimension in df.columns
                and dimension2 in df.columns
            ):
                tmp = df[[dimension, dimension2, metric]].dropna().copy()
                tmp[dimension] = tmp[dimension].astype(str)
                tmp[dimension2] = tmp[dimension2].astype(str)
                tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
                tmp = tmp.dropna(subset=[metric])
                gg = tmp.groupby([dimension, dimension2])[metric].sum(numeric_only=True).reset_index()
                gg = gg.sort_values(by=metric, ascending=False).head(50)
                lines = []
                for _, r in gg.iterrows():
                    s = str(r[dimension])
                    t = str(r[dimension2])
                    cs = name_to_centroid.get(s)
                    ct = name_to_centroid.get(t)
                    if not cs or not ct:
                        continue
                    lines.append(
                        {
                            "fromName": s,
                            "toName": t,
                            "coords": [[cs[0], cs[1]], [ct[0], ct[1]]],
                            "value": float(r[metric] or 0),
                        }
                    )
                if lines:
                    series_out.append(
                        {
                            "type": "lines",
                            "coordinateSystem": "geo",
                            "zlevel": 2,
                            "effect": {"show": True, "symbol": "arrow", "symbolSize": 8},
                            "lineStyle": {"width": 1.2, "opacity": 0.45, "curveness": 0.2, "color": "#2FA3FF"},
                            "data": lines,
                        }
                    )

            series_out = [s for s in series_out if s]
            for s in series_out:
                if s.get("rippleEffect") is None:
                    s.pop("rippleEffect", None)

            option = {
                "title": {"text": f"{metric} 地图", "textStyle": {"color": "#1277D1", "fontSize": 16}},
                "tooltip": {"trigger": "item"},
                "geo": base_geo,
                "visualMap": {
                    "min": 0,
                    "max": vmax,
                    "left": "left",
                    "bottom": 0,
                    "inRange": {"color": ["#CCE5FF", "#2FA3FF", "#1277D1"]},
                    "textStyle": {"color": "#3A4A63"},
                    "calculable": True,
                    "show": style == "区域填色",
                },
                "series": series_out,
            }
            return ChartBundle(
                title=f"{metric} 地图分布",
                table=g,
                echarts_option=option,
                echarts_extra={"registerMap": {"name": map_name, "geoJSON": map_geojson}},
            )

        if lon_col and lat_col and lon_col != "不选择" and lat_col != "不选择" and lon_col in df.columns and lat_col in df.columns:
            tmp = df[[lon_col, lat_col, metric] + ([dimension] if dimension and dimension != "不选择" and dimension in df.columns else [])].dropna().copy()
            tmp[lon_col] = pd.to_numeric(tmp[lon_col], errors="coerce")
            tmp[lat_col] = pd.to_numeric(tmp[lat_col], errors="coerce")
            tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
            tmp = tmp.dropna(subset=[lon_col, lat_col, metric]).head(3000)
            data = tmp[[lon_col, lat_col, metric]].values.tolist()
            option = {
                "title": {"text": f"{metric} 坐标分布", "textStyle": {"color": "#1277D1", "fontSize": 16}},
                "tooltip": {"trigger": "item"},
                "xAxis": {"type": "value", "name": lon_col, "axisLabel": {"color": "#3A4A63"}},
                "yAxis": {"type": "value", "name": lat_col, "axisLabel": {"color": "#3A4A63"}},
                "series": [{"type": "scatter", "data": data, "symbolSize": 8, "itemStyle": {"color": "#1277D1", "opacity": 0.75}}],
            }
            return ChartBundle(title=f"{metric} 坐标分布", table=tmp, echarts_option=option)

        return None

    return None


def _plot_trend(df: pd.DataFrame, metrics: list[str], x_axis_mode: str, style: Optional[str]) -> ChartBundle:
    x, x_name = _extract_x(df, x_axis_mode)
    use_cols = [m for m in metrics if m in df.columns]
    tmp = df.copy()
    tmp["__x"] = x
    agg = tmp.groupby("__x")[use_cols].sum(numeric_only=True).reset_index()

    fig = go.Figure()
    for m in use_cols:
        if str(style) == "面积":
            fig.add_trace(go.Scatter(x=agg["__x"], y=agg[m], mode="lines", name=m, fill="tozeroy"))
        else:
            fig.add_trace(go.Scatter(x=agg["__x"], y=agg[m], mode="lines", name=m))

    fig.update_layout(
        title="趋势分析",
        xaxis_title=x_name,
        yaxis_title="数值",
        legend_title_text="指标",
        margin=dict(l=20, r=20, t=50, b=20),
        template="plotly_white",
    )

    table = agg.rename(columns={"__x": x_name})
    title = " · ".join(use_cols) + " 趋势" if use_cols else "趋势图"
    return ChartBundle(title=title, figure=fig, table=table)


def _plot_pie(df: pd.DataFrame, metric: str, dimension: str, style: Optional[str]) -> ChartBundle:
    g = df.groupby(dimension)[metric].sum(numeric_only=True).reset_index()
    g = g.sort_values(by=metric, ascending=False).head(30)
    title = f"{metric} 按 {dimension} 占比"

    if str(style) == "玫瑰":
        fig = go.Figure(
            data=[
                go.Barpolar(r=g[metric].tolist(), theta=g[dimension].astype(str).tolist(), name=metric)
            ]
        )
        fig.update_layout(
            title=title,
            polar=dict(radialaxis=dict(showticklabels=False, ticks="")),
            margin=dict(l=20, r=20, t=50, b=20),
        )
    else:
        fig = go.Figure(
            data=[go.Pie(labels=g[dimension].astype(str), values=g[metric], hole=0.55, name=metric)]
        )
        fig.update_layout(title=title, margin=dict(l=20, r=20, t=50, b=20))

    return ChartBundle(title=title, figure=fig, table=g)


def _plot_bar(df: pd.DataFrame, metric: str, dimension: str, style: Optional[str]) -> ChartBundle:
    g = df.groupby(dimension)[metric].sum(numeric_only=True).reset_index()
    g = g.sort_values(by=metric, ascending=False).head(10)
    title = f"{metric} 按 {dimension} 排名"

    if str(style) == "竖向":
        fig = go.Figure(data=[go.Bar(x=g[dimension].astype(str), y=g[metric], name=metric)])
        fig.update_layout(title=title, xaxis_title=dimension, yaxis_title=metric, margin=dict(l=20, r=20, t=50, b=20))
    else:
        fig = go.Figure(data=[go.Bar(y=g[dimension].astype(str), x=g[metric], orientation="h", name=metric)])
        fig.update_layout(title=title, xaxis_title=metric, yaxis_title=dimension, margin=dict(l=20, r=20, t=50, b=20))
        fig.update_yaxes(autorange="reversed")

    return ChartBundle(title=title, figure=fig, table=g)


def _plot_funnel(df: pd.DataFrame, metrics: list[str], dimension: str) -> ChartBundle:
    title = "转化漏斗"
    if dimension and dimension != "不选择" and metrics:
        metric = metrics[0]
        g = df.groupby(dimension)[metric].sum(numeric_only=True).reset_index().sort_values(by=metric, ascending=False)
        g = g.head(20)
        fig = go.Figure(data=[go.Funnel(y=g[dimension].astype(str), x=g[metric])])
        title = f"{metric} 按 {dimension} 漏斗"
        fig.update_layout(title=title, margin=dict(l=20, r=20, t=50, b=20))
        return ChartBundle(title=title, figure=fig, table=g)

    use_cols = [m for m in metrics if m in df.columns]
    sums = df[use_cols].sum(numeric_only=True)
    g = pd.DataFrame({"stage": sums.index.tolist(), "value": sums.values.tolist()})
    fig = go.Figure(data=[go.Funnel(y=g["stage"], x=g["value"])])
    fig.update_layout(title="多指标转化漏斗", margin=dict(l=20, r=20, t=50, b=20))
    title = " → ".join(use_cols) + " 转化漏斗" if use_cols else "漏斗图"
    return ChartBundle(title=title, figure=fig, table=g)


def _plot_boxplot(df: pd.DataFrame, metric: str, dimension: str) -> ChartBundle:
    tmp = df[[dimension, metric]].dropna().copy()
    tmp[dimension] = tmp[dimension].astype(str)
    medians = tmp.groupby(dimension)[metric].median().sort_values(ascending=False)
    cats = medians.index.tolist()[:20]
    tmp = tmp[tmp[dimension].isin(cats)]

    fig = go.Figure()
    for c in cats:
        vals = tmp[tmp[dimension] == c][metric]
        fig.add_trace(go.Box(y=vals, name=c, boxmean=True))
    title = f"{metric} 按 {dimension} 分布（箱线）"
    fig.update_layout(title=title, margin=dict(l=20, r=20, t=50, b=20))
    return ChartBundle(title=title, figure=fig, table=tmp)


def build_chart_bundles(df: pd.DataFrame, chart_confs: list[dict[str, Any]]) -> list[ChartBundle]:
    bundles: list[ChartBundle] = []
    for conf in chart_confs:
        ctype = conf.get("type")
        metrics = conf.get("metrics") or []
        dimension = conf.get("dimension")
        x_axis_mode = conf.get("x_axis_mode")
        style = conf.get("chart_style")
        dimension2 = str(conf.get("dimension2") or "不选择")
        target_value = float(conf.get("target_value") or 0)
        agg_mode = str(conf.get("agg_mode") or "汇总")
        radar_topn = int(conf.get("radar_topn") or 5)
        map_geojson = conf.get("map_geojson")
        lon_col = str(conf.get("lon_col") or "不选择")
        lat_col = str(conf.get("lat_col") or "不选择")

        echarts_bundle = _build_echarts_bundle(df, conf)

        plotly_figure: Optional[go.Figure] = None
        plotly_table: Optional[pd.DataFrame] = None
        if ctype == "趋势图" and metrics:
            b = _plot_trend(df, metrics, str(x_axis_mode), style)
            plotly_figure, plotly_table = b.figure, b.table
        elif ctype == "占比图" and metrics and dimension and dimension != "不选择":
            b = _plot_pie(df, metrics[0], str(dimension), style)
            plotly_figure, plotly_table = b.figure, b.table
        elif ctype == "排名图" and metrics and dimension and dimension != "不选择":
            b = _plot_bar(df, metrics[0], str(dimension), style)
            plotly_figure, plotly_table = b.figure, b.table
        elif ctype == "漏斗图" and metrics:
            b = _plot_funnel(df, metrics, str(dimension))
            plotly_figure, plotly_table = b.figure, b.table
        elif ctype == "箱线图" and metrics and dimension and dimension != "不选择":
            b = _plot_boxplot(df, metrics[0], str(dimension))
            plotly_figure, plotly_table = b.figure, b.table
        elif ctype == "散点图" and len(metrics) >= 2:
            x_metric, y_metric = metrics[0], metrics[1]
            tmp = df[[x_metric, y_metric] + ([dimension] if dimension and dimension != "不选择" and dimension in df.columns else [])].dropna().copy()
            tmp[x_metric] = pd.to_numeric(tmp[x_metric], errors="coerce")
            tmp[y_metric] = pd.to_numeric(tmp[y_metric], errors="coerce")
            tmp = tmp.dropna(subset=[x_metric, y_metric]).head(2000)
            fig = go.Figure()
            if dimension and dimension != "不选择" and dimension in tmp.columns:
                top_cats = tmp[dimension].astype(str).value_counts().head(8).index.tolist()
                for c in top_cats:
                    sub = tmp[tmp[dimension].astype(str) == c]
                    fig.add_trace(go.Scatter(x=sub[x_metric], y=sub[y_metric], mode="markers", name=str(c)))
            else:
                fig.add_trace(go.Scatter(x=tmp[x_metric], y=tmp[y_metric], mode="markers", name=f"{x_metric} vs {y_metric}"))
            fig.update_layout(title=f"{x_metric} vs {y_metric}", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
            plotly_figure, plotly_table = fig, tmp
        elif ctype == "热力图" and metrics and dimension and dimension != "不选择" and dimension2 != "不选择":
            metric = metrics[0]
            tmp = df[[dimension, dimension2, metric]].dropna().copy()
            tmp[dimension] = tmp[dimension].astype(str)
            tmp[dimension2] = tmp[dimension2].astype(str)
            pivot = tmp.groupby([dimension2, dimension])[metric].sum(numeric_only=True).reset_index()
            mat = pivot.pivot(index=dimension2, columns=dimension, values=metric).fillna(0)
            fig = go.Figure(data=go.Heatmap(z=mat.values, x=mat.columns.astype(str), y=mat.index.astype(str), colorscale=[[0, "#CCE5FF"], [0.5, "#2FA3FF"], [1, "#1277D1"]]))
            fig.update_layout(title=f"{metric} 热力分布", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
            plotly_figure, plotly_table = fig, pivot
        elif ctype == "雷达图" and metrics and dimension and dimension != "不选择":
            use_metrics = [m for m in metrics[:6] if m in df.columns]
            tmp = df[[dimension] + use_metrics].dropna().copy()
            tmp[dimension] = tmp[dimension].astype(str)
            g = tmp.groupby(dimension)[use_metrics].sum(numeric_only=True).reset_index()
            sort_metric = use_metrics[0] if use_metrics else None
            if sort_metric:
                g = g.sort_values(by=sort_metric, ascending=False).head(max(1, int(radar_topn or 5)))
            fig = go.Figure()
            for _, row in g.iterrows():
                fig.add_trace(go.Scatterpolar(r=[float(row[m] or 0) for m in use_metrics], theta=use_metrics, fill="toself", name=str(row[dimension])))
            fig.update_layout(title="多维指标对比", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
            plotly_figure, plotly_table = fig, g
        elif ctype == "仪表盘" and metrics:
            metric = metrics[0]
            series = pd.to_numeric(df[metric], errors="coerce") if metric in df.columns else pd.Series([], dtype=float)
            v = float(series.mean() if str(agg_mode) == "均值" else series.sum()) if len(series) else 0.0
            if pd.isna(v):
                v = 0.0
            max_v = float(target_value or 0)
            if max_v <= 0:
                max_v = max(1.0, abs(v) * 1.2)
            fig = go.Figure(go.Indicator(mode="gauge+number", value=v, gauge={"axis": {"range": [0, max_v]}, "bar": {"color": "#1277D1"}}))
            fig.update_layout(title=f"{metric} 完成度", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
            plotly_figure, plotly_table = fig, pd.DataFrame({"metric": [metric], "value": [v], "target": [max_v], "agg_mode": [agg_mode]})
        elif ctype == "桑基图" and metrics and dimension and dimension != "不选择" and dimension2 != "不选择":
            metric = metrics[0]
            tmp = df[[dimension, dimension2, metric]].dropna().copy()
            tmp[dimension] = tmp[dimension].astype(str)
            tmp[dimension2] = tmp[dimension2].astype(str)
            tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
            tmp = tmp.dropna(subset=[metric])
            g = tmp.groupby([dimension, dimension2])[metric].sum(numeric_only=True).reset_index()
            g = g.sort_values(by=metric, ascending=False).head(40)
            nodes = sorted(set(g[dimension].tolist() + g[dimension2].tolist()))
            node_index = {n: i for i, n in enumerate(nodes)}
            fig = go.Figure(
                data=[
                    go.Sankey(
                        node={"label": nodes, "pad": 10, "thickness": 14},
                        link={
                            "source": [node_index[str(s)] for s in g[dimension].tolist()],
                            "target": [node_index[str(t)] for t in g[dimension2].tolist()],
                            "value": [float(v or 0) for v in g[metric].tolist()],
                        },
                    )
                ]
            )
            fig.update_layout(title=f"{metric} 流向", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
            plotly_figure, plotly_table = fig, g
        elif ctype == "地图" and metrics:
            metric = metrics[0]
            if (
                lon_col
                and lat_col
                and lon_col != "不选择"
                and lat_col != "不选择"
                and lon_col in df.columns
                and lat_col in df.columns
                and metric in df.columns
            ):
                tmp = df[[lon_col, lat_col, metric]].dropna().copy()
                tmp[lon_col] = pd.to_numeric(tmp[lon_col], errors="coerce")
                tmp[lat_col] = pd.to_numeric(tmp[lat_col], errors="coerce")
                tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
                tmp = tmp.dropna(subset=[lon_col, lat_col, metric]).head(3000)
                fig = go.Figure(data=[go.Scatter(x=tmp[lon_col], y=tmp[lat_col], mode="markers", marker={"size": 8, "color": "#1277D1"})])
                fig.update_layout(title=f"{metric} 坐标分布", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
                plotly_figure, plotly_table = fig, tmp
            elif isinstance(map_geojson, dict) and map_geojson.get("type") == "FeatureCollection" and map_geojson.get("features"):
                features = map_geojson.get("features") or []
                name_to_centroid: dict[str, tuple[float, float]] = {}
                for f in features:
                    n = _geojson_feature_name(f)
                    c = _geojson_feature_centroid(f)
                    if n and c:
                        name_to_centroid[n] = c
                if dimension and dimension != "不选择" and dimension in df.columns and metric in df.columns:
                    tmp = df[[dimension, metric]].dropna().copy()
                    tmp[dimension] = tmp[dimension].astype(str)
                    tmp[metric] = pd.to_numeric(tmp[metric], errors="coerce")
                    tmp = tmp.dropna(subset=[metric])
                    g = tmp.groupby(dimension)[metric].sum(numeric_only=True).reset_index().sort_values(by=metric, ascending=False)
                    xs, ys, sizes, names = [], [], [], []
                    for _, r in g.head(80).iterrows():
                        name = str(r[dimension])
                        c = name_to_centroid.get(name)
                        if not c:
                            continue
                        xs.append(c[0])
                        ys.append(c[1])
                        sizes.append(float(r[metric] or 0))
                        names.append(name)
                    fig = go.Figure(data=[go.Scatter(x=xs, y=ys, mode="markers+text", text=names, textposition="top center", marker={"size": 8, "color": "#1277D1"})])
                    fig.update_layout(title=f"{metric} 地图分布", template="plotly_white", margin=dict(l=20, r=20, t=50, b=20))
                    plotly_figure, plotly_table = fig, g

        title = echarts_bundle.title if echarts_bundle else str(ctype or "图表")
        table = echarts_bundle.table if echarts_bundle else (plotly_table if plotly_table is not None else pd.DataFrame())
        bundles.append(
            ChartBundle(
                title=title,
                table=table,
                figure=plotly_figure,
                echarts_option=echarts_bundle.echarts_option if echarts_bundle else None,
                echarts_extra=echarts_bundle.echarts_extra if echarts_bundle else None,
            )
        )

    return bundles


_ECHARTS_JS_CACHE: Optional[str] = None


def _load_echarts_runtime_js() -> str:
    global _ECHARTS_JS_CACHE
    if _ECHARTS_JS_CACHE is not None:
        return _ECHARTS_JS_CACHE

    import streamlit_echarts
    from pathlib import Path

    root = Path(streamlit_echarts.__file__).resolve().parent
    js_candidates = list(root.rglob("*.js"))

    preferred = [p for p in js_candidates if "echarts" in p.name.lower() and p.is_file()]
    ordered = sorted(preferred or js_candidates, key=lambda p: p.stat().st_size if p.exists() else 0, reverse=True)

    for p in ordered[:80]:
        try:
            raw = p.read_bytes()
            head = raw[:500_000].decode("utf-8", errors="ignore")
            if "echarts" not in head.lower():
                continue
            js = raw.decode("utf-8", errors="ignore")
            if "echarts" in js:
                _ECHARTS_JS_CACHE = js
                return js
        except Exception:
            continue

    raise RuntimeError("无法加载 ECharts 运行时，请确认已安装 streamlit-echarts")


def export_dashboard_html(
    *,
    product_name: str,
    user_email: str,
    language_default: Literal["zh", "en"],
    cards: list[CardSummary],
    charts: list[ChartBundle],
) -> ExportResult:
    use_echarts = any(bool(c.echarts_option) for c in charts)
    use_plotly = any(bool(c.figure) and not c.echarts_option for c in charts)
    echarts_js = _load_echarts_runtime_js() if use_echarts else ""
    plotly_js = get_plotlyjs() if use_plotly else ""
    created_at = _now_utc_iso()

    i18n = {
        "zh": {
            "title": "已保存看板导出",
            "cards": "指标卡",
            "charts": "已保存图表",
            "total": "总计",
            "avg": "均值",
            "lang": "语言",
        },
        "en": {
            "title": "Saved Dashboard Export",
            "cards": "Metric Cards",
            "charts": "Saved Charts",
            "total": "Total",
            "avg": "Avg",
            "lang": "Language",
        },
    }

    def t(key: str) -> str:
        return i18n.get(language_default, i18n["zh"]).get(key, key)

    card_html = []
    for c in cards:
        card_html.append(
            f"""
<div class=\"card\">
  <div class=\"card-name\">{c.name}</div>
  <div class=\"card-value\">{format_number(c.total)}</div>
  <div class=\"card-sub\">{t('avg')}: {format_number(c.average)}</div>
</div>
"""
        )

    chart_sections = []
    echarts_payload = []
    plotly_sections = []
    for idx, bundle in enumerate(charts):
        if bundle.echarts_option is not None:
            div_id = f"echarts_{idx}_{uuid.uuid4().hex}"
            chart_sections.append(
                f"""
<section class=\"chart\">
  <div class=\"chart-title\">{bundle.title}</div>
  <div class=\"chart-body\"><div class=\"echarts\" id=\"{div_id}\"></div></div>
</section>
"""
            )
            echarts_payload.append(
                {
                    "id": div_id,
                    "option": bundle.echarts_option,
                    "extra": bundle.echarts_extra,
                }
            )
            continue

        if bundle.figure is None:
            continue
        div_id = f"plotly_{idx}_{uuid.uuid4().hex}"
        fig_html = bundle.figure.to_html(
            include_plotlyjs=False,
            full_html=False,
            config={"responsive": True, "displaylogo": False},
            div_id=div_id,
        )
        plotly_sections.append(
            f"""
<section class=\"chart\">
  <div class=\"chart-title\">{bundle.title}</div>
  <div class=\"chart-body\">{fig_html}</div>
</section>
"""
        )

    chart_html = "".join(chart_sections + plotly_sections)

    html = f"""<!doctype html>
<html lang=\"{language_default}\">
<head>
  <meta charset=\"utf-8\"/>
  <meta name=\"viewport\" content=\"width=device-width,initial-scale=1\"/>
  <title>{t('title')}</title>
  <style>
    :root {{
      --bg:#E6F3FF;
      --card-bg:rgba(255,255,255,.95);
      --primary:#1277D1;
      --text:#0B1B33;
      --sub:#3A4A63;
      --border:#A7C9FF;
    }}
    body {{
      margin:0;
      font-family:-apple-system,BlinkMacSystemFont,"Segoe UI","PingFang SC","Microsoft YaHei",Arial,sans-serif;
      background:var(--bg);
      color:var(--text);
    }}
    header {{
      padding:16px 20px;
      border-bottom:1px solid var(--border);
      background:linear-gradient(90deg, rgba(18,119,209,.12) 0%, transparent 100%);
      display:flex;
      align-items:center;
      justify-content:space-between;
      gap:12px;
    }}
    .h-title {{
      font-weight:700;
      letter-spacing:.6px;
      color:var(--primary);
    }}
    .lang {{
      display:flex;
      align-items:center;
      gap:8px;
      color:var(--sub);
      font-size:14px;
    }}
    main {{
      padding:18px 20px 40px;
    }}
    .cards {{
      display:grid;
      grid-template-columns:repeat(auto-fit, minmax(220px, 1fr));
      gap:12px;
      margin-bottom:18px;
    }}
    .card {{
      background:var(--card-bg);
      border:1px solid var(--border);
      padding:12px 14px;
      box-shadow: inset 0 0 18px rgba(18,119,209,.08), 0 8px 20px rgba(12,35,66,.06);
      clip-path: polygon(0 0, 100% 0, 100% calc(100% - 14px), calc(100% - 14px) 100%, 0 100%);
    }}
    .card-name {{
      color:var(--sub);
      font-size:14px;
    }}
    .card-value {{
      color:var(--primary);
      font-weight:800;
      font-size:30px;
      margin-top:6px;
    }}
    .card-sub {{
      color:var(--sub);
      font-size:13px;
      margin-top:4px;
    }}
    .chart {{
      background:var(--card-bg);
      border:1px solid var(--border);
      padding:12px 14px;
      margin-bottom:14px;
    }}
    .chart-title {{
      font-weight:700;
      margin-bottom:10px;
    }}
    .watermark {{
      position:fixed;
      right:16px;
      bottom:12px;
      opacity:.45;
      color:var(--sub);
      font-size:12px;
      pointer-events:none;
      user-select:none;
      background:rgba(255,255,255,.55);
      border:1px solid rgba(167,201,255,.8);
      padding:6px 8px;
      border-radius:8px;
      backdrop-filter: blur(6px);
    }}
  </style>
  <script>
    window.PlotlyConfig = {{MathJaxConfig: 'local'}};
  </script>
  <script>{plotly_js}</script>
  <script>{echarts_js}</script>
</head>
<body>
  <header>
    <div class=\"h-title\">{t('title')}</div>
    <div class=\"lang\">
      <span id=\"langLabel\">{t('lang')}</span>
      <select id=\"langSelect\" aria-label=\"language\">
        <option value=\"zh\">中文</option>
        <option value=\"en\">English</option>
      </select>
    </div>
  </header>
  <main>
    <section>
      <h3 data-i18n=\"cards\" style=\"margin:0 0 10px;color:var(--text)\">{t('cards')}</h3>
      <div class=\"cards\">{''.join(card_html)}</div>
    </section>
    <section>
      <h3 data-i18n=\"charts\" style=\"margin:0 0 10px;color:var(--text)\">{t('charts')}</h3>
      {chart_html}
    </section>
  </main>
  <div class="watermark">Generated by {product_name} | {created_at}</div>
  <script>
    const i18n = {json.dumps(i18n, ensure_ascii=False)};
    const select = document.getElementById('langSelect');
    select.value = document.documentElement.lang || '{language_default}';
    function applyLang(lang) {{
      document.documentElement.lang = lang;
      const dict = i18n[lang] || i18n['zh'];
      document.title = dict.title || document.title;
      document.getElementById('langLabel').innerText = dict.lang || 'Language';
      document.querySelectorAll('[data-i18n]').forEach((el) => {{
        const key = el.getAttribute('data-i18n');
        if (dict[key]) el.innerText = dict[key];
      }});
    }}
    select.addEventListener('change', (e) => applyLang(e.target.value));
    applyLang(select.value);
  </script>
  <script>
    const echartsCharts = {json.dumps(echarts_payload, ensure_ascii=False)};
    function initECharts() {{
      if (!echartsCharts || !echartsCharts.length) return;
      if (typeof echarts === 'undefined') return;
      echartsCharts.forEach((c) => {{
        try {{
          if (c.extra && c.extra.registerMap && c.extra.registerMap.name && c.extra.registerMap.geoJSON) {{
            echarts.registerMap(c.extra.registerMap.name, c.extra.registerMap.geoJSON);
          }}
          const el = document.getElementById(c.id);
          if (!el) return;
          const chart = echarts.init(el, null, {{renderer: 'canvas'}});
          chart.setOption(c.option || {{}}, true);
          window.addEventListener('resize', () => chart.resize());
        }} catch (e) {{
        }}
      }});
    }}
    initECharts();
  </script>
</body>
</html>"""

    data = html.encode("utf-8")
    return ExportResult(content=data, sha256=_sha256(data))


def _write_df_table(ws: Worksheet, df: pd.DataFrame, start_row: int, start_col: int) -> None:
    for j, col in enumerate(df.columns, start=start_col):
        cell = ws.cell(row=start_row, column=j, value=str(col))
        cell.font = Font(bold=True)
        cell.alignment = Alignment(horizontal="center")
    for i, (_, row) in enumerate(df.iterrows(), start=start_row + 1):
        for j, col in enumerate(df.columns, start=start_col):
            v = row[col]
            c = ws.cell(row=i, column=j, value=v)
            if isinstance(v, (int, float)) and not pd.isna(v):
                c.number_format = "0.0000"

    ws.freeze_panes = ws.cell(row=start_row + 1, column=start_col)
    ws.auto_filter.ref = f"{get_column_letter(start_col)}{start_row}:{get_column_letter(start_col + len(df.columns) - 1)}{start_row + len(df)}"

    for j, col in enumerate(df.columns, start=start_col):
        max_len = max([len(str(col))] + [len(str(v)) for v in df[col].head(100).tolist()])
        ws.column_dimensions[get_column_letter(j)].width = min(max(10, max_len + 2), 38)


def _echarts_option_to_png_bytes(
    *,
    option: dict[str, Any],
    extra: Optional[dict[str, Any]],
    width: int,
    height: int,
    pixel_ratio: int,
    timeout_ms: int,
) -> bytes:
    echarts_js = _load_echarts_runtime_js()
    option_json = json.dumps(option or {}, ensure_ascii=False)
    extra_json = json.dumps(extra or {}, ensure_ascii=False)
    html = f"""<!doctype html>
<html>
<head>
  <meta charset=\"utf-8\"/>
  <meta name=\"viewport\" content=\"width=device-width,initial-scale=1\"/>
  <style>
    html, body {{ margin:0; padding:0; background:#ffffff; }}
    #c {{ width:{width}px; height:{height}px; }}
  </style>
  <script>{echarts_js}</script>
</head>
<body>
  <div id=\"c\"></div>
  <script>
    window.__ECHARTS_READY__ = false;
    window.__ECHARTS_PNG__ = null;
    (function() {{
      try {{
        const option = {option_json};
        const extra = {extra_json};
        if (extra && extra.registerMap && extra.registerMap.name && extra.registerMap.geoJSON) {{
          echarts.registerMap(extra.registerMap.name, extra.registerMap.geoJSON);
        }}
        const el = document.getElementById('c');
        const chart = echarts.init(el, null, {{ renderer: 'canvas' }});
        chart.setOption(option || {{}}, true);
        window.__ECHARTS_READY__ = true;
        setTimeout(() => {{
          try {{
            window.__ECHARTS_PNG__ = chart.getDataURL({{ type: 'png', pixelRatio: {pixel_ratio}, backgroundColor: '#ffffff' }});
          }} catch (e) {{
            window.__ECHARTS_PNG__ = null;
          }}
        }}, 250);
      }} catch (e) {{
        window.__ECHARTS_READY__ = false;
      }}
    }})();
  </script>
</body>
</html>"""

    def try_playwright() -> Optional[bytes]:
        try:
            from playwright.sync_api import sync_playwright
        except Exception:
            return None

        tmp_path = None
        try:
            with tempfile.NamedTemporaryFile("w", suffix=".html", delete=False, encoding="utf-8") as f:
                f.write(html)
                tmp_path = f.name

            with sync_playwright() as p:
                browser = p.chromium.launch()
                page = browser.new_page(viewport={"width": int(width), "height": int(height)})
                page.goto("file:///" + tmp_path.replace("\\", "/"), wait_until="load")
                page.wait_for_function(
                    "window.__ECHARTS_PNG__ && window.__ECHARTS_PNG__.startsWith('data:image/png;base64,')",
                    timeout=timeout_ms,
                )
                data_url = page.evaluate("window.__ECHARTS_PNG__")
                browser.close()
            if not isinstance(data_url, str) or "base64," not in data_url:
                return None
            b64 = data_url.split("base64,", 1)[1]
            return base64.b64decode(b64)
        except Exception:
            return None
        finally:
            if tmp_path:
                try:
                    import os

                    os.unlink(tmp_path)
                except Exception:
                    pass

    png = try_playwright()
    if png:
        return png

    raise RuntimeError("ECharts 图表截图生成失败：缺少可用的浏览器渲染环境")


def export_dashboard_xlsx(
    *,
    product_name: str,
    user_email: str,
    cards: list[CardSummary],
    charts: list[ChartBundle],
) -> ExportResult:
    wb = Workbook()
    wb.remove(wb.active)

    stamp = f"Generated by {product_name} {_now_utc_iso()}"

    for c in cards:
        ws = wb.create_sheet(_safe_sheet_title(f"Dashboard_{c.name}"))
        ws["A1"].value = stamp
        ws["A1"].font = Font(color="666666")
        ws["A3"].value = "Metric"
        ws["B3"].value = c.name
        ws["A4"].value = "Total"
        ws["B4"].value = c.total
        ws["A5"].value = "Avg"
        ws["B5"].value = c.average
        ws["A3"].font = Font(bold=True)
        ws["A4"].font = Font(bold=True)
        ws["A5"].font = Font(bold=True)
        ws.column_dimensions["A"].width = 16
        ws.column_dimensions["B"].width = 24
        ws.freeze_panes = "A4"

        for r in (4, 5):
            if isinstance(ws[f"B{r}"].value, (int, float)) and ws[f"B{r}"].value is not None:
                ws[f"B{r}"].number_format = "0.0000"

    if charts:
        ws = wb.create_sheet(_safe_sheet_title("Dashboard_Charts"))
        ws["A1"].value = stamp
        ws["A1"].font = Font(color="666666")
        row_cursor = 3
        for idx, bundle in enumerate(charts, start=1):
            ws.cell(row=row_cursor, column=1, value=f"{idx}. {bundle.title}").font = Font(bold=True)
            row_cursor += 1

            img_bytes: Optional[bytes] = None
            if bundle.echarts_option is not None:
                try:
                    img_bytes = _echarts_option_to_png_bytes(
                        option=bundle.echarts_option,
                        extra=bundle.echarts_extra,
                        width=1400,
                        height=720,
                        pixel_ratio=2,
                        timeout_ms=7_000,
                    )
                except Exception:
                    img_bytes = None
            if img_bytes is None and bundle.figure is not None:
                img_bytes = bundle.figure.to_image(format="png", width=1400, height=720, scale=2)
            if img_bytes is not None:
                img_bio = io.BytesIO(img_bytes)
                xl_img = XLImage(img_bio)
                img_cell = f"A{row_cursor}"
                ws.add_image(xl_img, img_cell)

            table_start_col = 14
            _write_df_table(ws, bundle.table.reset_index(drop=True), row_cursor, table_start_col)

            row_cursor += max(24, len(bundle.table) + 6)

    bio = io.BytesIO()
    wb.save(bio)
    data = bio.getvalue()
    return ExportResult(content=data, sha256=_sha256(data))


class ExportTaskManager:
    def __init__(self, max_workers: int = 4):
        self._executor = ThreadPoolExecutor(max_workers=max_workers)
        self._lock = threading.Lock()
        self._tasks: dict[str, ExportTask] = {}

    def get(self, task_id: str) -> Optional[ExportTask]:
        with self._lock:
            return self._tasks.get(task_id)

    def cancel(self, task_id: str, *, user_email: str) -> bool:
        with self._lock:
            task = self._tasks.get(task_id)
            if not task or task.user_email != user_email:
                return False
            task.cancelled = True
            if task._future and task._future.cancel():
                task.status = "cancelled"
                task.progress = 100
                task.finished_at_utc = _now_utc_iso()
            return True

    def get_result(self, task_id: str, *, user_email: str) -> Optional[ExportResult]:
        with self._lock:
            task = self._tasks.get(task_id)
            if not task or task.user_email != user_email or not task._future:
                return None
            future = task._future
        if not future.done():
            return None
        if future.cancelled():
            return None
        exc = future.exception()
        if exc is not None:
            return None
        return future.result()

    def submit(
        self,
        *,
        user_email: str,
        export_format: ExportFormat,
        estimate: ExportEstimate,
        runner: Callable[[Callable[[int], None], Callable[[], bool]], ExportResult],
    ) -> ExportTask:
        task_id = str(uuid.uuid4())
        trace_id = str(uuid.uuid4())
        task = ExportTask(
            task_id=task_id,
            user_email=user_email,
            export_format=export_format,
            estimate_seconds=estimate.estimate_seconds,
            size_bytes=estimate.size_bytes,
            trace_id=trace_id,
        )

        def set_progress(p: int) -> None:
            with self._lock:
                t = self._tasks.get(task_id)
                if not t:
                    return
                t.progress = max(0, min(100, int(p)))

        def is_cancelled() -> bool:
            with self._lock:
                t = self._tasks.get(task_id)
                return bool(t and t.cancelled)

        def job() -> ExportResult:
            with self._lock:
                t = self._tasks.get(task_id)
                if t:
                    t.status = "running"
                    t.progress = 1
            try:
                result = runner(set_progress, is_cancelled)
                with self._lock:
                    t = self._tasks.get(task_id)
                    if t and t.cancelled:
                        t.status = "cancelled"
                    elif t:
                        t.status = "succeeded"
                        t.sha256 = result.sha256
                    if t:
                        t.progress = 100
                        t.finished_at_utc = _now_utc_iso()
                return result
            except Exception as e:
                with self._lock:
                    t = self._tasks.get(task_id)
                    if t:
                        t.status = "cancelled" if t.cancelled else "failed"
                        t.error_message = str(e)
                        t.progress = 100
                        t.finished_at_utc = _now_utc_iso()
                raise

        with self._lock:
            self._tasks[task_id] = task

        future = self._executor.submit(job)
        with self._lock:
            self._tasks[task_id]._future = future

        return task


TASK_MANAGER = ExportTaskManager()
